#General library imports
from typing import Union, Optional, List
import base64
import os
import datetime
from datetime import date, timedelta
import csv
import aiofiles
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy


#Import FastAPI dependencies
from fastapi import Depends, FastAPI , HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi_sqlalchemy import DBSessionMiddleware,db
from fastapi.responses import FileResponse

#Import SQLAlchemy dependencies
from sqlalchemy import func, desc, text, case, Date, and_, or_
from auth_utils import get_user

#Import libraries for token decryption
from decimal import Decimal

#Import schemas from schemas.py
from schemas import HoursUserBase as SchemaHoursUserBase
from schemas import BufferDailyRegister as SchemaBufferDailyRegister
from schemas import FrontEndUser as SchemaFrontEndUser
from schemas import Project as SchemaProject
from schemas import Record as SchemaRecord
from schemas import RecordProjects as SchemaRecordProjects
from schemas import Favorites as SchemaFavorites
from schemas import FrontendProjectPhase as SchemaProjectPhase
from schemas import MonthlyModifiedHours as SchemaMonthlyModifiedHours
from schemas import ProjectMonthlyInformation as SchemaProjectMonthlyInformation
from schemas import MonthlyReport as SchemaMonthlyReport
from schemas import MonthlyModifiedItems as SchemaMonthlyModifiedItems
#Import models from models.py
from models import HoursUser as ModelHoursUser
from models import Project as ModelProject
from models import Record as ModelRecord
from models import RecordProjects as ModelRecordProjects
from models import Favorites as ModelFavorites
from models import BufferDailyRegister as ModelBufferDailyRegister
from models import ProjectPhase as ModelProjectPhase
from models import MonthlyModifiedHours as ModelMonthlyModifiedHours
from models import ProjectMonthlyInformation as ModelProjectMonthlyInformation
from models import MonthlyReport as ModelMonthlyReport

from database import SessionLocal
from dotenv import load_dotenv


load_dotenv(".env")


app = FastAPI()


def get_db():
    db=SessionLocal()
    try:
        yield db
    finally:
        db.close()



# Middlewares
origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


app.add_middleware(DBSessionMiddleware, db_url=os.environ["DATABASE_URL"])


def wrap_jwk_public_key(jwk):
    # Extract the "n" (modulus) and "e" (exponent) from the JWK
    modulus = base64.urlsafe_b64decode(jwk["n"] + "==")
    exponent = base64.urlsafe_b64decode(jwk["e"] + "==")

    # Construct the RSA public key in ASN.1 format
    public_key_asn1 = (
        b"\x30\x0d\x06\x09\x2a\x86\x48\x86\xf7\x0d\x01\x01\x01\x05\x00\x03"
        + b"\x81\x81\x00"
        + modulus
        + b"\x02\x03\x01\x00\x01\x03\x81\x81\x00"
        + exponent
    )

    # Convert the public key to Base64-encoded string
    public_key_base64 = base64.b64encode(public_key_asn1).decode("utf-8")

    # Wrap the Base64-encoded key with BEGIN and END headers
    
    wrapped_key = " ".join([public_key_base64[i:i+64] for i in range(0, len(public_key_base64), 64)])
    

    return wrapped_key




#Root message
@app.get("/")
def read_root():
    return {"status": "OK"}


# PART 1: METHODS FOR THE EMPLOYEE PROFILE
#____________________________________________________________________________________________________


# 1.1 Get user
@app.get("/api/user")
async def get_hours_user(user: SchemaHoursUserBase = Depends(get_user)):
    
    if not user:
            raise HTTPException(status_code=400, detail="User not found")
    model_user=db.session.query(ModelHoursUser).filter(ModelHoursUser.username  == user.username,
                                                     ModelHoursUser.email  == user.email).first()

    frontenduser = SchemaFrontEndUser(
        email=model_user.email,
        username=model_user.username,
        id=model_user.id,  
        domain=model_user.domain,
        role=model_user.role,
        view=model_user.view,
        date_entrance=model_user.date_entrance,
        status=model_user.status)

    return frontenduser

# 1.2 Get user ID 
@app.get("/api/userID")
async def get_user_ID(user: SchemaHoursUserBase = Depends(get_user)):
    
    if not user:
            raise HTTPException(status_code=400, detail="User not found")
    
    usermodel=db.session.query(ModelHoursUser).filter(ModelHoursUser.email  == user.email).first()
    
    return str(usermodel.id)


# 1.3 Insert record
@app.post("/api/records")
async def insert_record(record: SchemaRecord, record_projects: List[SchemaRecordProjects], user: SchemaHoursUserBase = Depends(get_user)):
    
    userID = await get_user_ID(user)

    #Before everything: verify the user
    if not str(record.user_id) == userID:
        raise HTTPException(status_code=401, detail="Unauthorized access")
    
    if record.date_rec.weekday() != 2:
            raise HTTPException(status_code=400,detail="Invalid date")

    #First part: defining class record
    if not db.session.query(ModelHoursUser).filter(ModelHoursUser.id == record.user_id).first():
            raise HTTPException(status_code=400, detail="User not found")
        
    existing_record = db.session.query(ModelRecord).filter(
        ModelRecord.user_id == record.user_id,
        ModelRecord.date_rec == record.date_rec).first()
           
    if existing_record:
        raise HTTPException(status_code=409, detail="Record already exists")

    db_record = ModelRecord(
            user_id = record.user_id,
            comment = record.comment,
            date_rec = record.date_rec
        )
    
    #Second part: adding the project to the created record
    db_projects = []
    hourcount = 0.0
    
    for record_project in record_projects:
      
        if not project_exists(record_project.project_id):
            raise HTTPException(status_code=404,detail="Project not found")
        
        existing_project = db.session.query(ModelRecordProjects).filter(
            ModelRecordProjects.user_id == record.user_id,
            ModelRecordProjects.date_rec == record.date_rec,
            ModelRecordProjects.project_id == record_project.project_id
        ).first()

        if existing_project:
            raise HTTPException(status_code=409, detail="User has already filled hours for this project in this week")

        if not record_project.declared_hours == 0.0:
            db_project = ModelRecordProjects(
                user_id = record.user_id,
                date_rec = record.date_rec,
                project_id = record_project.project_id,
                declared_hours = record_project.declared_hours,
                domain = record_project.domain
            )
            db_projects.append(db_project)

            searchedPhase=db.session.query(ModelProjectPhase).filter(
                ModelProjectPhase.project_id == record_project.project_id,
                ModelProjectPhase.start_date <= record.date_rec,
                ModelProjectPhase.end_date >= record.date_rec).first()
            
            if searchedPhase:
                searchedPhase.hours += Decimal(record_project.declared_hours)

            
        
        hourcount += record_project.declared_hours
   
    if hourcount != 35.0:
            raise HTTPException(status_code=400,detail="Hour count does not match required value")
    
    date_init = record.date_rec - datetime.timedelta(days=4)

    searched_records = db.session.query(ModelBufferDailyRegister).filter(
        ModelBufferDailyRegister.user_id == record.user_id,
        ModelBufferDailyRegister.day_date.between(date_init, record.date_rec)
        ).all()
    
    for buffer_record in searched_records:
        db.session.delete(buffer_record)

 
    db.session.add_all(db_projects)
    db.session.add(db_record)
    db.session.commit()
    return {"message": "Record created successfully."}


# 1.4 Get record
@app.get("/api/records/{hours_user_id}")
async def get_records(hours_user_id: int, user: SchemaHoursUserBase = Depends(get_user)):
    
    ans = []

    if not db.session.query(ModelHoursUser).filter(ModelHoursUser.id == hours_user_id).first():
        raise HTTPException(status_code=400, detail="User not found")
    
    userID = await get_user_ID(user)
    
    if not str(hours_user_id) == userID:
        raise HTTPException(status_code=401, detail="Unauthorized access")

    user_declaredrecords = db.session.query(ModelRecord).filter(
        ModelRecord.user_id == hours_user_id).order_by(desc(ModelRecord.date_rec)).all()
    
    for record in user_declaredrecords:
        rec_projects = []

        record_projects = db.session.query(ModelRecordProjects).filter(ModelRecordProjects.user_id == hours_user_id,
                                                                       ModelRecordProjects.date_rec == record.date_rec,
                                                                       ModelRecordProjects.declared_hours != 0,
                                                                       ).all()

        for project in record_projects:
            rec_proj = SchemaRecordProjects(
                project_id= project.project_id,
                declared_hours= project.declared_hours,
                domain=project.domain
            )
            rec_projects.append(rec_proj)
        
        rec_ans = {"record": record,
                   "record_projects":rec_projects}
        ans.append(rec_ans)     

    return ans


#1.5 Modify hours in records
@app.put("/api/records")
async def change_record(record:SchemaRecord, record_projects:List[SchemaRecordProjects], user: SchemaHoursUserBase = Depends(get_user)):

    hourcount=0.0

    current_record_projects = db.session.query(ModelRecordProjects).filter(
            ModelRecordProjects.user_id == record.user_id,
            ModelRecordProjects.date_rec == record.date_rec,
        )
    
    for record_project in current_record_projects:
        db.session.delete(record_project)

    for rp in record_projects:

        db_rp = ModelRecordProjects(
            user_id = record.user_id,
            date_rec = record.date_rec,
            project_id = rp.project_id,
            declared_hours = rp.declared_hours,
            domain = rp.domain
            )
        db.session.add(db_rp)

        hourcount += rp.declared_hours
    
    if hourcount != 35.0:
            raise HTTPException(status_code=400,detail="Hour count does not match required value")


    db.session.commit()
    return {"message":"Change successful"}


# 1.6 Get projects /*/
@app.get("/api/projects",response_model=List[SchemaProject])
async def get_projects(user = Depends(get_user)):
    # THE FOLLOWING LINE IS FOCAL NAIM SPECIFIC, if you want to extract the project KPI software to other organizations, please implement the function : 
    #   filter projects which organization (entity) is one of the organization of the user. For Focal Naim, the name of organization is in email
    
    if (user.role == 'Business Manager'):
        projects =db.session.query(ModelProject).all()
    else:
        org = []
        if ('focal' in user.email):
            org.append('FOCAL')
        if ('naim' in user.email):
            org.append('NAIM')
        projects =db.session.query(ModelProject).filter(ModelProject.entity.in_(org)).all()
    if not projects:
            raise HTTPException(status_code=404,detail="No projects on database")
    return projects

# 1.7 Define favorites /*/
@app.post("/api/favorites",response_model=List[SchemaFavorites])
async def define_favorites(favorites: List[SchemaFavorites], user = Depends(get_user)):
    db_favorites = []
    
    for favorite in favorites:
        if not project_exists(favorite.project_id):
            raise HTTPException(status_code=404,detail="Project not found")
        
        if not db.session.query(ModelHoursUser).filter(ModelHoursUser.id == favorite.user_id).first():
            raise HTTPException(status_code=400, detail="User not found")

        existing_favorite = db.session.query(ModelFavorites).filter(
            ModelFavorites.project_id == favorite.project_id,
            ModelFavorites.user_id == favorite.user_id,
        ).first()
           
        if existing_favorite:
            raise HTTPException(status_code=409, detail="Favorite already exists")

        db_favorite = ModelFavorites(
            project_id=favorite.project_id,
            user_id=favorite.user_id
        )

        db_favorites.append(db_favorite)

    db.session.add_all(db_favorites)
    db.session.commit()
    return db_favorites

# 1.8 Get favorites /*/
@app.get("/api/favorites/{hours_user_id}",response_model=List[int])
async def get_favorites(hours_user_id: int , user = Depends(get_user)):

    if not db.session.query(ModelHoursUser).filter(ModelHoursUser.id == hours_user_id).first():
        raise HTTPException(status_code=400, detail="User not found")

    favorites =db.session.query(ModelFavorites.project_id).filter(
        ModelFavorites.user_id == hours_user_id
    ).all()

    projects_ids = [project_id for (project_id,) in favorites]
    return projects_ids

# 1.9 Delete favorites /*/
@app.delete("/api/favorites")
async def delete_favorite(favorite: SchemaFavorites, user = Depends(get_user)):
    user_id = favorite.user_id
    project_id = favorite.project_id

    if not db.session.query(ModelHoursUser).filter(ModelHoursUser.id == user_id).first():
        raise HTTPException(status_code=400, detail="User not found")
    
    if not db.session.query(ModelProject).filter(ModelProject.id == project_id).first():
        raise HTTPException(status_code=404, detail="Project not found")

    db_favorite = db.session.query(ModelFavorites).filter(
        ModelFavorites.user_id == user_id,
        ModelFavorites.project_id == project_id,
    ).first()

    if not db_favorite:
        raise HTTPException(status_code=404, detail="Project not defined as favorite by user")
    
    db.session.delete(db_favorite)
    db.session.commit()
    return {"message": "Favorite deleted"}

# 1.10 Get user domain /*/
@app.get("/api/domain/{hours_user_id}",response_model=str)
async def get_domain(hours_user_id: int, user = Depends(get_user)):
    searched_user = db.session.query(ModelHoursUser).filter(ModelHoursUser.id == hours_user_id).first()
    
    if not searched_user:
        raise HTTPException(status_code=400, detail="User not found")
        
    searched_user_domain = str(searched_user.domain)    
    return searched_user_domain

# 1.11 Update user domain /*/
@app.put("/api/domain/{hours_user_id}")
async def update_domain(hours_user_id: int, updated_domain: str, user = Depends(get_user)):
    
    domain_options = ['Acoustics','Tests','Hardware','Software','Mechanics','Project Management','Shop floor','General']

    if updated_domain not in domain_options:
        raise HTTPException(status_code=400, detail="Wrong domain input")
    
    searched_user = db.session.query(ModelHoursUser).get(hours_user_id)
    
    if not searched_user:
        raise HTTPException(status_code=400, detail="User not found")
        
    searched_user.domain = str(updated_domain)
    db.session.commit()
    return {"message": "Domain updated"}


# 1.12 Input to buffer table /*/
@app.post("/api/buffertable")
async def post_hours_per_day(dailyregs: List[SchemaBufferDailyRegister], user = Depends(get_user)):
    
    for dailyreg in dailyregs:
        searched_records = db.session.query(ModelBufferDailyRegister).filter(
            ModelBufferDailyRegister.user_id == dailyreg.user_id,
            ModelBufferDailyRegister.day_date == dailyreg.day_date,
            ).all()
        
        for buffer_record in searched_records:
            db.session.delete(buffer_record)
    
    for dailyreg in dailyregs:
        if dailyreg.daily_hours > 0:
            buffer_rec = ModelBufferDailyRegister(
                user_id = dailyreg.user_id,
                day_date = dailyreg.day_date,
                project_id = dailyreg.project_id,
                daily_hours = dailyreg.daily_hours
            )

            db.session.add(buffer_rec)
        
        
    
    db.session.commit()
    return {"status":"buffer days registered"}


# 1.13 Get buffer table records /*/
@app.get("/api/buffertable")
async def get_hours_per_day(hours_user_id: int, date_init: date, date_end: date, user = Depends(get_user)):
    searched_registers = db.session.query(ModelBufferDailyRegister).filter(
        ModelBufferDailyRegister.user_id == hours_user_id,
        ModelBufferDailyRegister.day_date.between(date_init, date_end)
        ).all()

    
    return searched_registers

# PART 2: METHODS FOR THE PROJECT MANAGER PROFILE
#____________________________________________________________________________________________________

# 2.1 Export records CSV 
@app.post("/api/export-records-csv")
async def export_csv(month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        projects: Optional[List[int]] = None):
    
    tempFileName='./temp/export.csv'

    if (month1 is None and year1 is not None) or (month1 is not None and year1 is None) or (month2 is None and year2 is not None) or (month2 is not None and year2 is None):
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    date_query = db.session.query(ModelRecord.date_rec).order_by(ModelRecord.date_rec)
    dateList=[row.date_rec for row in date_query]
    first_date=dateList[0]
    last_date=dateList[-1]


    projects_query = db.session.query(ModelProject.id)
    projectList=[row.id for row in projects_query]
    
    if not projects:
        projects = projectList


    if month1 and year1:
        dateM=datetime.date(year1,month1,1)
        if month2 and year2:
            dateFin=datetime.date(year2,month2+1,1)
        else:
            dateFin=last_date
    elif month2 and year2:
        dateFin=datetime.date(year2,month2+1,1)
        dateM=first_date
    else:
        dateM=first_date
        dateFin=last_date

    result = db.session.query(ModelRecordProjects.id.label('id'),
                        ModelProject.project_code.label('project_code'),
                        ModelRecordProjects.date_rec.label('date'),
                        ModelHoursUser.username.label('name'),
                        ModelHoursUser.email.label('email'),
                        ModelRecordProjects.domain.label('domain'),
                           ModelRecordProjects.declared_hours.label('hours'))\
    .join(ModelProject, ModelProject.id == ModelRecordProjects.project_id,isouter=True)\
    .join(ModelHoursUser, ModelHoursUser.id == ModelRecordProjects.user_id,isouter=True)\
    .filter(ModelRecordProjects.date_rec>=dateM,
            ModelRecordProjects.date_rec<=dateFin,
            ModelProject.id.in_(projects))\
    .order_by(ModelProject.project_code,ModelRecordProjects.date_rec,ModelHoursUser.username)

    column_names = ["project_code",
                    "week",
                    "year",
                    "name",
                    "email",
                    "domain",
                    "hours"]

    csv_data = csv.DictWriter(open(tempFileName,"w"), fieldnames=column_names,delimiter=";")
    csv_data.writeheader()
    
    for row in result:
       
        week=row.date.isocalendar()[1]
        year=row.date.year
        
        row_data = {"project_code":row.project_code,
            "week":week,
            "year":year,
            "name":row.name,
            "email":row.email,
            "domain":row.domain,
            "hours":row.hours}      
        csv_data.writerow(row_data)
        

    async with aiofiles.open(tempFileName, mode="r") as f:
        contents = await f.read()


    return FileResponse(tempFileName, filename=tempFileName)

# 2.2 Import records CSV
@app.post("/api/import-csv")
async def import_csv(file: UploadFile = File(...)):
    
    

    try:

        contents = await file.read()
        csv_data = csv.DictReader(contents.decode("utf-8").splitlines(),delimiter=";")
        hoursRecords={}

        for row in csv_data:
            project_code = str(row['project_code'])
            week = int(row['week'])
            year = int(row['year'])
            name = str(row['name'])
            email = str(row['email'])
            domain = str(row['domain'])
            hours = float(row['hours'])

            

            date_rec = datetime.date.fromisocalendar(year,week,3)
            
            searched_user = db.session.query(ModelHoursUser).filter(
                                            ModelHoursUser.email == email).first()
            
            if not searched_user:
                raise HTTPException(status_code=400, detail="User "+ name + " not found")
            
            user_id = getattr(searched_user,"id")
            
            searched_project = db.session.query(ModelProject).filter(
                                            ModelProject.project_code == project_code).first()
            
            if not searched_project:
                raise HTTPException(status_code=400, detail="Project "+ project_code + " not found")
            

            project_id = getattr(searched_project,"id")
            
            
            
            
            recordkey = str(date_rec)+str(user_id)



            record_exists = db.session.query(ModelRecord).filter(
                                            ModelRecord.user_id == user_id,
                                            ModelRecord.date_rec == date_rec).first()

                        
                
            if not record_exists:
                new_record = ModelRecord(
                user_id = user_id,
                date_rec = date_rec,
                )
                    
                db.session.add(new_record)
                hoursRecords[recordkey]=hours

            else:
                if recordkey not in hoursRecords:
                    
                    existing_record_projects=db.session.query(ModelRecordProjects).filter(
                                            ModelRecordProjects.user_id == user_id,
                                            ModelRecordProjects.date_rec == date_rec).all()
                    for rp in existing_record_projects:
                        db.session.delete(rp)
                    hoursRecords[recordkey]=hours
                else:
                    hoursRecords[recordkey]+=hours

            
            new_record_project = ModelRecordProjects(
                        user_id = user_id,
                        date_rec = date_rec,
                        project_id = project_id,
                        domain = domain,
                        declared_hours = hours)
            
                
                
            db.session.add(new_record_project)
        
        all_lower = all(value == 35 for value in hoursRecords.values())
        
        if not all_lower:
            raise HTTPException(status_code=400, detail="Some lines do not sum 35 hours per week. Please check your input")
        
        
        db.session.commit()
        return {"message": "Import successful."}
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid input, please try again")


    

    
# 2.3 Add project
@app.post("/api/project")
async def add_project(project: SchemaProject, phases: List[SchemaProjectPhase], project_monthly_information: List[SchemaProjectMonthlyInformation], user = Depends(get_user)):

    existing_code = db.session.query(ModelProject).filter(ModelProject.project_code == project.project_code).first() 
    
    if existing_code:
        raise HTTPException(status_code=400, detail="Project code already exists")
    
    db_project = ModelProject(
        entity = project.entity,
        division = project.division,
        sub_category = project.sub_category,
        classification = project.classification,
        type = project.type,
        project_name = project.project_name,
        project_code = project.project_code,
        project_manager = project.project_manager,
        complexity = project.complexity,
        start_cap_date = project.start_cap_date,
        end_cap_date = project.end_cap_date,
        start_date = project.start_date,
        end_date = project.end_date,
        status = project.status
        )
    
    db.session.add(db_project)
    
    projectid = getattr(db.session.query(ModelProject).filter(ModelProject.project_code == project.project_code).first(),"id")

    for phase in phases:
     
        db_phase = ModelProjectPhase(
            project_id = projectid,
            project_phase = phase.project_phase,
            start_date = phase.start_date,
            end_date = phase.end_date
            )
        
        db.session.add(db_phase)
    
    for info in project_monthly_information:
        if info.month is not None and info.year is not None:
            p_info_date = datetime.date(info.year, info.month, 1)
        else:
            raise HTTPException(status_code=400, detail="Invalid date") 

        db_project_monthly_info = ModelProjectMonthlyInformation(
            project_id = projectid,
            month = p_info_date,
            forecast_hours = info.forecast_hours,
            capitalizable = info.capitalizable
            )
        db.session.add(db_project_monthly_info)

    db.session.commit()
    
    return {'message: Project with phases added successfully'}

# 2.3 Change project state
@app.put("/api/project/change_state")
async def change_project_state(project_id: int, status: str):

    edited_project = db.session.query(ModelProject).filter(ModelProject.id == project_id).first() 
    if not edited_project:
        raise HTTPException(status_code=400, detail="Project id doesn't exist")
    edited_project.status = status
    db.session.commit()

    return {'message: Project state successfully updated'}

getprojectmodel={'project': SchemaProject,
            "phases":List[SchemaProjectPhase]} 

# 2.4 Get project /*/
@app.get("/api/project")
async def get_project_with_phases(projectcode: str, user = Depends(get_user)):
    
    modelproject =db.session.query(ModelProject).filter(
        ModelProject.project_code == projectcode
    ).first()

    if not modelproject:
            raise HTTPException(status_code=404,detail="Project does not exist")
    
    project = SchemaProject(
        id = modelproject.id,
        entity = modelproject.entity,
        division = modelproject.division,
        sub_category= modelproject.sub_category,
        classification= modelproject.classification, 
        type= modelproject.type,
        project_name= modelproject.project_name,
        project_code= modelproject.project_code,
        project_manager= modelproject.project_manager,
        complexity= modelproject.complexity,
        start_cap_date = modelproject.start_cap_date,
        end_cap_date = modelproject.end_cap_date,
        start_date = modelproject.start_date,
        end_date = modelproject.end_date,
        status = modelproject.status,
    )

    phases_ans=[]

    projectid=getattr(project,"id")

    phases = db.session.query(ModelProjectPhase).filter(
        ModelProjectPhase.project_id == projectid
    )

    if phases:
        for phase in phases:
            frontend_phase=SchemaProjectPhase(
                project_phase = phase.project_phase, 
                start_date = phase.start_date,
                end_date = phase.end_date
            )
            phases_ans.append(frontend_phase)

    p_monthly_infos_ans=[]

    p_monthly_infos = db.session.query(ModelProjectMonthlyInformation).filter(
        ModelProjectMonthlyInformation.project_id == projectid
    )
    
    if p_monthly_infos:
        for p_monthly_info in p_monthly_infos:
            frontend_p_monthly_info=SchemaProjectMonthlyInformation(
                project_id = projectid,
                month = p_monthly_info.month.month,
                year = p_monthly_info.month.year,
                forecast_hours = p_monthly_info.forecast_hours,
                capitalizable = p_monthly_info.capitalizable,
            )
            p_monthly_infos_ans.append(frontend_p_monthly_info)


    return {'project': project,
            "phases":phases_ans,
            "monthly_informations":p_monthly_infos_ans} 



# 2.5 Modify project
@app.put("/api/project")
async def update_project(project: SchemaProject, phases: List[SchemaProjectPhase], monthly_informations: List[SchemaProjectMonthlyInformation], user = Depends(get_user)):
           
    searched_project = db.session.query(ModelProject).filter(ModelProject.id == project.id).first()

    existing_code = db.session.query(ModelProject).filter(ModelProject.project_code == project.project_code).first() 
    
    if existing_code:
        if not (existing_code.project_code == searched_project.project_code):
            raise HTTPException(status_code=400, detail="Project code already exists")
    
    if not searched_project:
        raise HTTPException(status_code=400, detail="Project not found")
    
    searched_project.entity = project.entity
    searched_project.division = project.division
    searched_project.sub_category = project.sub_category
    searched_project.classification = project.classification
    searched_project.type = project.type
    searched_project.project_name = project.project_name
    searched_project.project_code = project.project_code
    searched_project.project_manager = project.project_manager
    searched_project.complexity = project.complexity
    searched_project.start_cap_date = project.start_cap_date
    searched_project.end_cap_date = project.end_cap_date
    searched_project.start_date = project.start_date
    searched_project.end_date = project.end_date
        
    db.session.commit()


    searched_phases = db.session.query(ModelProjectPhase).filter(
        ModelProjectPhase.project_id == project.id)

    for searched_phase in searched_phases:
        db.session.delete(searched_phase)
    
    db.session.commit()
    
    for phase in phases:

        db_phase = ModelProjectPhase(
            project_id = project.id,
            project_phase = phase.project_phase,
            start_date = phase.start_date,
            end_date = phase.end_date
            )
    
        db.session.add(db_phase)
    
    db.session.commit()

    update_project_phase_hours(project.id)

    searched_p_m_infos = db.session.query(ModelProjectMonthlyInformation).filter(
        ModelProjectMonthlyInformation.project_id == project.id)
    
    for searched_p_m_info in searched_p_m_infos:
        db.session.delete(searched_p_m_info)

    for p_m_infos in monthly_informations:

        if p_m_infos.month is not None and p_m_infos.year is not None:
            p_m_infos_date = datetime.date(p_m_infos.year, p_m_infos.month, 1)
        else:
            raise HTTPException(status_code=400, detail="Invalid date") 

        db_p_m_info = ModelProjectMonthlyInformation(
            project_id = project.id,
            month = p_m_infos_date,
            hours = p_m_infos.forecast_hours,
            capitalizable = p_m_infos.capitalizable
            )
    
        db.session.add(db_p_m_info)
    
    db.session.commit()


    return {"message": "Project updated"}


# 2.6 Delete project
@app.delete("/api/project")
async def delete_project(project_code: str, user: SchemaHoursUserBase = Depends(get_user)):
      
    project_id = getattr(db.session.query(ModelProject).filter(ModelProject.project_code==project_code).first(),"id")
    
    if not validateManagers(False,user):
        raise HTTPException(status_code=401, detail="Unauthorized access")
    
    searched_project = db.session.query(ModelProject).filter(ModelProject.id == project_id).first()
    
    if not searched_project:
        raise HTTPException(status_code=400, detail="Project not found")
    
    message="Project not deleted. Please erase the records associated to the project first."
    
    records_in_project=db.session.query(ModelRecordProjects).filter(
        ModelRecordProjects.project_id == project_id 
    ).first()
        
    if not records_in_project:
        
        to_delete_phases = db.session.query(ModelProjectPhase).filter(ModelProjectPhase.project_id == project_id).all()
        to_delete_p_m_infos = db.session.query(ModelProjectMonthlyInformation).filter(ModelProjectMonthlyInformation.project_id == project_id).all()
        
        for phase in to_delete_phases:
            db.session.delete(phase)
            db.session.commit()

        for p_m_info in to_delete_p_m_infos:
            db.session.delete(p_m_info)
            db.session.commit()

        db.session.delete(searched_project)
        message="Project deleted."
        db.session.commit()


    
        
   
    return {"message": message}

# KPIs

# 2.7. KPI pie chart
@app.get("/api/kpi/pie/hours_by_domain")
async def kpi_piedomainhours( project_id: int, 
        unit: str, 
        month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        user: SchemaHoursUserBase = Depends(get_user)
    ):
    
    start_date = None
    end_date = None

    if month1 is not None and year1 is not None:
        start_date = datetime.date(year1, month1, 1)

    if month2 is not None and year2 is not None:
        if month2 == 12:
            end_date = datetime.date(year2 + 1, 1, 1)
        else:
            end_date = datetime.date(year2, month2 + 1, 1)


    if start_date:
        if end_date:
            result = (db.session.query(ModelRecordProjects.domain,func.sum(ModelRecordProjects.declared_hours))
              .filter(
            ModelRecordProjects.project_id == project_id,
            ModelRecordProjects.date_rec >= start_date,
            ModelRecordProjects.date_rec < end_date
            ).group_by(ModelRecordProjects.domain))
        else:
            (db.session.query(ModelRecordProjects.domain,func.sum(ModelRecordProjects.declared_hours))
              .filter(
            ModelRecordProjects.project_id == project_id,
            ModelRecordProjects.date_rec >= start_date,
            )
            .group_by(ModelRecordProjects.domain))
    elif end_date:
        result = (db.session.query(ModelRecordProjects.domain,func.sum(ModelRecordProjects.declared_hours))
              .filter(
            ModelRecordProjects.project_id == project_id,
            ModelRecordProjects.date_rec < end_date
            )
            .group_by(ModelRecordProjects.domain))
    else:
        result = (db.session.query(ModelRecordProjects.domain,func.sum(ModelRecordProjects.declared_hours))
              .filter(
            ModelRecordProjects.project_id == project_id
            )
            .group_by(ModelRecordProjects.domain))
    
    
    
    
    factor = 1
    if(unit=='TDE'):
        factor=140
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")


    declared_hours_by_domain = {domain: round(declared_hours/factor,2) for domain, declared_hours in result}



    return {
        'unit':'h',
        'series':[
            {
                "data": declared_hours_by_domain,
                "name": "Spent hours by domain",
                "type": 'pie'
            }
        ]
    }

#2.8 KPI Line graph
@app.get("/api/kpi/line/hour_expenditure")
async def kpi_linehours(
        project_id: int, 
        cumulative: bool, 
        unit: str, 
        month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        user: SchemaHoursUserBase = Depends(get_user)
    ):
    
    # Get every month where there is info about forecast
    # The forecast hours must be a successive array of months
    hours_and_forecast_query = db.session.query(
        ModelProjectMonthlyInformation.month.label('month'),
        ModelProjectMonthlyInformation.forecast_hours.label('forecast_hours'),
        ModelMonthlyModifiedHours.total_hours.label('hours')
    ).outerjoin(ModelMonthlyModifiedHours, ModelMonthlyModifiedHours.month == ModelProjectMonthlyInformation.month
           ).filter(
            ModelProjectMonthlyInformation.project_id == project_id).all()

    # First date last date calculus
    months=[]
    hours=[]
    forecast_hours=[]
    for row in hours_and_forecast_query:
        hours.append(row.month)
        forecast_hours.append(row.forecast_hours)
        months.append(row.month)
    # Convert the result into a list of dictionaries    
    print(hours)
    converted_hours = hours.copy()
    converted_forecast_hours = forecast_hours.copy()

    if(unit=='TDE'):
        TDEs=[round(row/140,2) for row in hours]
        converted_hours=TDEs
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")

    if(unit=='TDE'):
        TDEs=[round(row/140,2) for row in forecast_hours]
        converted_forecast_hours=TDEs
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")
        

    if cumulative:
        cumhours = []
        cumsum=0.0
        for rhour in converted_hours:
            cumsum += float(rhour)
            cumhours.append(cumsum)
        converted_hours=cumhours

    if cumulative:
        cumforehours = []
        cumforesum=0.0
        for fhour in converted_forecast_hours:
            cumforesum += round(float(fhour),2)
            cumforehours.append(cumforesum)
        converted_forecast_hours=cumforehours


    converted_hours=[round(row,2) for row in converted_hours]
    converted_forecast_hours = [round(row,2) for row in converted_forecast_hours]

    # Format
    legend=["Spent"]    
    series= [
            {
                "data": converted_hours,
                "name": "Spent",
                "type": 'line'
            }
        ]    
    if (sum(forecast_hours)!=0):
        series.append({
            "data": converted_forecast_hours,
            "name": "Forecast",
            "type": 'line'
            }
            )
        legend.append("Forecast")
    ans= {
        'unit': unit,
        'xAxis': {"data": months},
        'series': series,
        'legend': {
            'data': legend
        }
    }

    return ans

#2.9 KPI Stacked bar chart
@app.get("/api/kpi/stackedbar/hour_expenditure_by_project")
async def kpi_stackedbar(        
        project_id: int,
        unit: str,
        user: SchemaHoursUserBase = Depends(get_user)
    ):

    
    stackedbar=[]
    phases=[]

    factor = 1
    if(unit=='TDE'):
        factor=140
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")
        
    result = (
        db.session.query(
            ModelProjectPhase.project_phase,ModelProjectPhase.hours
        ).filter(ModelProjectPhase.project_id == project_id).all()
        )

    for res in result:
                       
        to_add={
            "data": [round(res.hours/factor,2)],
            "name": res.project_phase,
            "type": "bar"
        }
        
        stackedbar.append(to_add)
        phases.append(res.project_phase)

    return {
        'unit':unit,
        'yAxis': {"data": ["Time spent"]},
        'series':stackedbar,
        'legend':{'data':phases}
    }

#2.10 See data
@app.post("/api/data")
async def see_data(month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        projects: Optional[List[int]] = None,
        user: SchemaHoursUserBase = Depends(get_user)):
    
    if (month1 is None and year1 is not None) or (month1 is not None and year1 is None) or (month2 is None and year2 is not None) or (month2 is not None and year2 is None):
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    date_query = db.session.query(ModelRecord.date_rec).order_by(ModelRecord.date_rec)
    dateList=[row.date_rec for row in date_query]
    first_date=dateList[0]
    last_date=dateList[-1]


    projects_query = db.session.query(ModelProject.id)
    projectList=[row.id for row in projects_query]
    
    if not projects:
        projects = projectList


    if month1 and year1:
        dateM=datetime.date(year1,month1,1)
        if month2 and year2:
            dateFin=datetime.date(year2,month2+1,1)
        else:
            dateFin=last_date
    elif month2 and year2:
        dateFin=datetime.date(year2,month2+1,1)
        dateM=first_date
    else:
        dateM=first_date
        dateFin=last_date
    
    ans=[]



    query = db.session.query(ModelRecordProjects.id.label('id'),
                        ModelProject.project_code.label('project_code'),
                        ModelRecordProjects.date_rec.label('date'),
                        ModelHoursUser.username.label('name'),
                        ModelHoursUser.email.label('email'),
                           ModelRecordProjects.declared_hours.label('hours'))\
    .join(ModelProject, ModelProject.id == ModelRecordProjects.project_id,isouter=True)\
    .join(ModelHoursUser, ModelHoursUser.id == ModelRecordProjects.user_id,isouter=True)\
    .filter(ModelRecordProjects.date_rec>=dateM,
            ModelRecordProjects.date_rec<=dateFin,
            ModelProject.id.in_(projects))\
    .order_by(ModelProject.project_code,ModelRecordProjects.date_rec,ModelHoursUser.username)

    for row in query:
        
        week=row.date.isocalendar()[1]
        year=row.date.year

        db_row={
            "project_code":row.project_code,
            "week":week,
            "year":year,
            "name":row.name,
            "email":row.email,
            "hours":row.hours,
            "id":row.id,
        }
        
        ans.append(db_row)

    return ans


#2.11 Export projects
@app.get("/api/projects/export")
async def get_projects():
    wb = openpyxl.Workbook()
    organizations = os.environ['ORGANIZATIONS'].split(',')
    widths = [3,13,13,13,13,13,40,10]
    for org in organizations:
        ws = wb.create_sheet(org) 
        modelprojects =db.session.query(ModelProject)
        ws.append([])
        ws.append([None,'Project Code','Division','Sub category','Classification','Expansion/Renewal','Project name','Status'])
        i = 0
        for modelproject in modelprojects:
            ws.append([None,modelproject.project_code,modelproject.division,modelproject.sub_category,modelproject.classification, modelproject.type, modelproject.project_name,modelproject.status])
            i += 1
        
        for j in range(0,len(widths)):
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(j+1)].width = widths[j]

        style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,showFirstColumn=False,
                       showLastColumn=False, showColumnStripes=False)
        tab = openpyxl.worksheet.table.Table(displayName=f"ProjectMasterList{org}", ref=f"B2:H{i+2}")
        tab.tableStyleInfo = style
        ws.add_table(tab)
    del wb['Sheet']
    temp_filename = "./temp/projects_export.xlsx"
    wb.save(filename=temp_filename)
    return FileResponse(temp_filename, filename=temp_filename)


# PART 3: METHODS FOR THE BUSINESS MANAGER PROFILE
#____________________________________________________________________________________________________

#3.1. Get monthly hours
@app.get("/api/monthlyhours",response_model=List[SchemaMonthlyModifiedHours])
async def get_monthly_hours(month:int, year:int, user: SchemaHoursUserBase = Depends(get_user)):
    
    
    response=[]
    # Get the current date and time
    
    dateM=datetime.date(year,month,1)

    # Get users who have filled there hours for the current month
    filtered_users=db.session.query(ModelMonthlyModifiedHours.user_id).filter(ModelMonthlyModifiedHours.month == dateM).distinct().all()
    
    for user in filtered_users:
        
        
        username = getattr(db.session.query(ModelHoursUser).filter(ModelHoursUser.id == user.user_id).first(),"username")
        tothours=0.0
        hourslist=[]
        dom="General"

        # Get monthly records
        records_with_user=db.session.query(ModelMonthlyModifiedHours).filter(ModelMonthlyModifiedHours.month == dateM,
                                                                             ModelMonthlyModifiedHours.user_id == user.user_id)

        # Get monthly records
        for record in records_with_user:
            if record.total_hours:
                tothours = record.total_hours
            if not record.total_hours == 0:
                proj={"project_id":record.project_id,
                    "hours":record.total_hours,
                    "domain": record.domain}
                hourslist.append(proj)

        if tothours != 0.0:
            sch=SchemaMonthlyModifiedHours(
                user_id=user.user_id,
                user_name=username,
                hours=hourslist
            )
            response.append(sch)

    return response


#3.2. Modify monthly hours
@app.put("/api/monthlyhours")
async def change_monthly_hours(month:int, year:int, changed_records:SchemaMonthlyModifiedItems, user: SchemaHoursUserBase = Depends(get_user)):
    
    dateM=datetime.date(year,month,1)

    for change in changed_records.hours_items:
        user_id = getattr(db.session.query(ModelMonthlyModifiedHours).filter(
                ModelMonthlyModifiedHours.user_id==change.user_id).first(),"user_id")


        for project in change.hours:
            searched_records=db.session.query(ModelMonthlyModifiedHours).filter(
                ModelMonthlyModifiedHours.project_id==project["project_id"],
                ModelMonthlyModifiedHours.month==dateM,
                ModelMonthlyModifiedHours.user_id==user_id)
            if searched_records:
                for rec in searched_records:
                    db.session.delete(rec)
            db_monthlyhour = ModelMonthlyModifiedHours(
                user_id = user_id,
                project_id = project["project_id"],
                month = dateM,
                total_hours = project["hours"],
                domain = project["domain"]
            )

            db.session.add(db_monthlyhour)
            
            db.session.commit()
    
    for m_i in changed_records.project_monthly_informations:
    
        searched_record=db.session.query(ModelProjectMonthlyInformation).filter(
            ModelProjectMonthlyInformation.project_id==m_i.project_id,
            ModelProjectMonthlyInformation.month==dateM).first()
        if searched_record:
            searched_record.capitalizable = m_i.capitalizable
        else:
            db_monthlyhour = ModelProjectMonthlyInformation(
                project_id = m_i.project_id,
                month = dateM,
                capitalizable = m_i.capitalizable,
            )

            db.session.add(db_monthlyhour)
        db.session.commit()

    return {"message":"Monthly Hours Project modified successfully"}


#3.3. Reset monthly hours' table
@app.post("/api/monthlyhours")
async def update_monthly_hours(month:int, year:int, user: SchemaHoursUserBase = Depends(get_user)):
    
    dateM=datetime.date(year,month,1)

    if month == 12:
        dateFin = datetime.date(year + 1, 1, 1)
    else:
        dateFin = datetime.date(year, month + 1, 1)


 
    existingMonthlyHours= db.session.query(ModelMonthlyModifiedHours).filter(
        ModelMonthlyModifiedHours.month == dateM
    ).all()

    monthly_report= db.session.query(ModelMonthlyReport).filter(
        ModelMonthlyReport.month == dateM
    ).first()
    print(monthly_report.sync_date)
    if (monthly_report):
        monthly_report.sync_date = datetime.datetime.now()
        db.session.commit()

    if existingMonthlyHours:
        for rec in existingMonthlyHours:
            db.session.delete(rec)
            db.session.commit()
    
    
    subquery = db.session.query(
        ModelRecordProjects.user_id,
        ModelRecordProjects.project_id,
        ModelRecordProjects.domain,
        func.date_trunc('month', dateM).label('month'),
        func.sum(
            case(
                (func.date_trunc('month', func.cast(
                                ModelRecordProjects.date_rec, 
                                Date
                            )) == dateM,ModelRecordProjects.declared_hours)),
                else_=0.0
            ).label('total_hours')
        ).filter(
            ModelRecordProjects.date_rec >= dateM,
            ModelRecordProjects.date_rec < dateFin,
            ModelRecordProjects.declared_hours != 0,
    ).group_by(
        ModelRecordProjects.user_id,
        ModelRecordProjects.project_id,
        ModelRecordProjects.domain,
        func.date_trunc('month', dateM)
    ).all()
    
    
    
    for row in subquery:
        
        db_monthlyhour = ModelMonthlyModifiedHours(
                user_id = row.user_id,
                project_id = row.project_id,
                month = dateM,
                total_hours = row.total_hours,
                domain = row.domain
                )
        db.session.add(db_monthlyhour)
        db.session.commit()

  
    return {"message":"Table updated successfully"}



#3.4. Business KPIs, capitalization summary pie

@app.get("/api/business_kpi/pie/cap_summary")
async def businesskpi_piecapsummary(
        unit: str, 
        month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        user: SchemaHoursUserBase = Depends(get_user)):
    
    if (month1 is None and year1 is not None) or (month1 is not None and year1 is None) or (month2 is None and year2 is not None) or (month2 is not None and year2 is None):
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    date_query = db.session.query(ModelMonthlyModifiedHours.month).order_by(ModelMonthlyModifiedHours.month)
    dateList=[row.month for row in date_query]
    first_date=dateList[0]
    last_date=dateList[-1]

    if month1 and year1:
        dateM=datetime.date(year1,month1,1)
        if month2 and year2:
            dateFin=datetime.date(year2,month2,1)
        else:
            dateFin=last_date
    elif month2 and year2:
        dateFin=datetime.date(year2,month2,1)
        dateM=first_date
    else:
        dateM=first_date
        dateFin=last_date
        
    #Capitalized
    
    cap_table_by_division= db.session.query(ModelProject.division, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month >= dateM,
            ModelMonthlyModifiedHours.month <= dateFin,
            or_(
                and_(
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= dateM,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > dateM)),
                and_( \
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= dateFin,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > dateM))))\
    .group_by(ModelProject.division) \
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())

    if not cap_table_by_division:
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    tot_cap_hours=sum([row.hours for row in cap_table_by_division])

    tot_hours_query= db.session.query(func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
    .filter(ModelMonthlyModifiedHours.month >= dateM,
            ModelMonthlyModifiedHours.month <= dateFin,
            ModelProject.sub_category != 'ABS').first()
    
    tot_hours = tot_hours_query.hours
    
    if not tot_hours:
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    tot_noncap_hours=tot_hours-tot_cap_hours


    factor = 1
    if(unit=='TDE'):
        factor=140
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")

    data1=[]
    
    db_cap_hours={"value": round(tot_cap_hours/factor,2),
            "name": 'Capitalized'}
    db_noncap_hours={"value": round(tot_noncap_hours/factor,2),
            "name": 'Non-capitalized'}
    
    data1.append(db_cap_hours)
    data1.append(db_noncap_hours)


    data2=[]
    for row in cap_table_by_division:
        db_cap={
            "value": round(row.hours/factor,2),
            "name": row.division
        }

        data2.append(db_cap)

    FCS_hours_query= db.session.query(func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
    .filter(ModelMonthlyModifiedHours.month >= dateM,
            ModelMonthlyModifiedHours.month <= dateFin,
            ModelProject.project_code == 'FCS').first()

    FCS_hours = FCS_hours_query.hours

    other_hours_query= db.session.query(func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
    .filter(ModelMonthlyModifiedHours.month >= dateM,
            ModelMonthlyModifiedHours.month <= dateFin,
            ModelProject.project_code != 'FCS',
            ModelProject.sub_category == 'ETC').first()

    other_hours = other_hours_query.hours

    noncap_proj_hours= tot_noncap_hours-FCS_hours-other_hours

    
    data2.append({"value":round(noncap_proj_hours/factor,2),
                 "name":"Non-cap projects"})
    
    data2.append({"value":round(FCS_hours/factor,2),
                 "name":"Eng. change/prod support"})
    
    data2.append({"value":round(other_hours/factor,2),
                 "name":"Man/Res/Reg/ISIT/Etc"})

    legend=['Capitalized','Non-capitalized']

    for row in cap_table_by_division:
        legend.append(row.division)

    legend.append("Non-cap projects")
    legend.append("Eng. change/prod support")
    legend.append("Man/Res/Reg/ISIT/Etc")

    ans={"legend":legend,
        "series":[
        {"type": "pie",
         "data":data1},
        {"type": "pie",
         "data":data2},
        ]
         }
    return ans




#3.5. Business KPIs, capitalization summary bar

@app.get("/api/business_kpi/bar/cap_summary")
async def businesskpi_barcapsummary(
        unit: str, 
        month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        user: SchemaHoursUserBase = Depends(get_user)):
    

    if (month1 is None and year1 is not None) or (month1 is not None and year1 is None) or (month2 is None and year2 is not None) or (month2 is not None and year2 is None):
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    date_query = db.session.query(ModelMonthlyModifiedHours.month).order_by(ModelMonthlyModifiedHours.month)
    dateList=[row.month for row in date_query]
    first_date=dateList[0]
    last_date=dateList[-1]

    if month1 and year1:
        dateM=datetime.date(year1,month1,1)
        if month2 and year2:
            dateFin=datetime.date(year2,month2,1)
        else:
            dateFin=last_date
    elif month2 and year2:
        dateFin=datetime.date(year2,month2,1)
        dateM=first_date
    else:
        dateM=first_date
        dateFin=last_date




    
    ans=[]

    cap_table_by_subcategory= db.session.query(ModelProject.sub_category, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month >= dateM,
            ModelMonthlyModifiedHours.month <= dateFin,
            or_(
                and_(
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= dateM,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > dateM)),
                and_( \
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= dateFin,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > dateM)))) \
    .group_by(ModelProject.sub_category) \
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())

    tot_table_by_subcategory= db.session.query(ModelProject.sub_category, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month >= dateM,
            ModelMonthlyModifiedHours.month <= dateFin,
            ModelProject.sub_category != 'ABS') \
    .group_by(ModelProject.sub_category) \
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())

    cap_subcats = [row.sub_category for row in cap_table_by_subcategory]
    tot_subcats = [row.sub_category for row in tot_table_by_subcategory]
    
    subcats=concatanateListswoduplicates(cap_subcats,tot_subcats)

    cap_dict={row.sub_category:row.hours for row in cap_table_by_subcategory}
    tot_dict={row.sub_category:row.hours for row in tot_table_by_subcategory}

    cap_values=[]
    tot_values=[]

    factor = 1
    if(unit=='TDE'):
        factor=140
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")
        
    for sc in subcats:
        cap_values.append(round(cap_dict.get(sc,0)/factor,2))
        tot_values.append(round(tot_dict.get(sc,0)/factor,2))

    ans={
        "yAxis":{'data': subcats},
        "series": [
            {"name":"Dev effort",
             "data":tot_values},
            {"name":"Capitalization",
             "data":cap_values}
        ]
    }

    return ans


#3.6. Business KPIs: stacked non-cap line

@app.get("/api/business_kpi/line/noncap_summary")
async def businesskpi_linenoncapsummary(
        unit: str, 
        month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        user: SchemaHoursUserBase = Depends(get_user)):

    if (month1 is None and year1 is not None) or (month1 is not None and year1 is None) or (month2 is None and year2 is not None) or (month2 is not None and year2 is None):
        raise HTTPException(status_code=400, detail="Invalid month and year input")

    date_query = db.session.query(ModelMonthlyModifiedHours.month).order_by(ModelMonthlyModifiedHours.month)
    dateList=[row.month for row in date_query]
    first_date=dateList[0]
    last_date=dateList[-1]

    if month1 and year1:
        dateM=datetime.date(year1,month1,1)
        if month2 and year2:
            dateFin=datetime.date(year2,month2,1)
        else:
            dateFin=last_date
    elif month2 and year2:
        dateFin=datetime.date(year2,month2,1)
        dateM=first_date
    else:
        dateM=first_date
        dateFin=last_date

    factor = 1
    if(unit=='TDE'):
        factor=140
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")
        
    

    FCS_hours_query = db.session.query(ModelMonthlyModifiedHours.month,func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month >= dateM,
                ModelMonthlyModifiedHours.month <= dateFin,
                ModelProject.project_code == 'FCS')\
        .group_by(ModelMonthlyModifiedHours.month)\
        .order_by(ModelMonthlyModifiedHours.month)
    
    ETC_hours_query = db.session.query(ModelMonthlyModifiedHours.month,func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month >= dateM,
                ModelMonthlyModifiedHours.month <= dateFin,
                ModelProject.sub_category == 'ETC',
                ModelProject.project_code != 'FCS')\
        .group_by(ModelMonthlyModifiedHours.month)\
        .order_by(ModelMonthlyModifiedHours.month)
    
    TOT_hours_query = db.session.query(ModelMonthlyModifiedHours.month,func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month >= dateM,
                ModelMonthlyModifiedHours.month <= dateFin,
                ModelProject.sub_category != 'ABS')\
        .group_by(ModelMonthlyModifiedHours.month)\
        .order_by(ModelMonthlyModifiedHours.month)
    
    dates=[row.month for row in FCS_hours_query]
    dataengchange=[round(row.hours/factor,2) for row in FCS_hours_query]
    
    dataetc=[round(row.hours/factor,2) for row in ETC_hours_query]
    datatot=[round(row.hours/factor,2) for row in TOT_hours_query]
    datacap=[]
    for date in dates:
        
        CAP_hours_query = db.session.query(ModelMonthlyModifiedHours.month,func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month == date,
                ModelProject.sub_category != 'ABS',
                and_(
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= date,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > date)))\
        .group_by(ModelMonthlyModifiedHours.month)\
        .order_by(ModelMonthlyModifiedHours.month)
        
        datacap.append(round(row.hours/factor,2) for row in CAP_hours_query)

    flatdatacap=[num for sublist in datacap for num in sublist]
    datacap=flatdatacap
    
    datanoncaptot=[a-b for a,b in zip(datatot,datacap)]
    datasum=[a+b for a,b in zip(dataetc,dataengchange)]
    datanoncap=[a-b for a,b in zip(datanoncaptot,datasum)]
    
    datatotal=[a+b for a,b in zip(dataengchange,datanoncap)]

    dates=[convert_date_to_month(date) for date in dates]

    ans = {
        'xAxis':{
            "data":dates
        },
        'series':[
            {
                'name': 'Non-cap projects',
                'data': datanoncap
            },
            {
                'name': 'Eng. change/prod support',
                'data': dataengchange
            },
            {
                'name': 'Total',
                'data': datatotal
            }
        ]
    }
    

    return ans

#3.7. Business KPIs, line
@app.get("/api/business_kpi/line/hour_expenditure")
async def businesskpi_linehours(
        cumulative:bool, 
        unit: str, 
        month1: Optional[int] = None,
        year1: Optional[int] = None,
        month2: Optional[int] = None,
        year2: Optional[int] = None,
        user: SchemaHoursUserBase = Depends(get_user)):
    
    
    start_date = None
    end_date = None

    if month1 is not None and year1 is not None:
        start_date = datetime.date(year1, month1, 1)

    if month2 is not None and year2 is not None:
        if month2 == 12:
            end_date = datetime.date(year2 + 1, 1, 1)
        else:
            end_date = datetime.date(year2, month2 + 1, 1)
        
    
    if start_date:
        if end_date:
            result = (
            db.session.query(
                ModelMonthlyModifiedHours.month.label('date'),
                func.sum(ModelMonthlyModifiedHours.total_hours).label('hours')
            )
            .filter(ModelMonthlyModifiedHours.month >= start_date,
                    ModelMonthlyModifiedHours.month < end_date)
            .group_by(ModelMonthlyModifiedHours.month)
            .order_by(ModelMonthlyModifiedHours.month)
            .all()
            )
        else: 
            result = (
            db.session.query(
                ModelMonthlyModifiedHours.month.label('date'),
                func.sum(ModelMonthlyModifiedHours.total_hours).label('hours')
            )
            .filter(ModelMonthlyModifiedHours.month >= start_date)
            .group_by(ModelMonthlyModifiedHours.month)
            .order_by(ModelMonthlyModifiedHours.month)
            .all()
            )
    elif end_date: result = (
            db.session.query(
                ModelMonthlyModifiedHours.month.label('date'),
                func.sum(ModelMonthlyModifiedHours.total_hours).label('hours')
            )
            .filter(ModelMonthlyModifiedHours.month < end_date)
            .group_by(ModelMonthlyModifiedHours.month)
            .order_by(ModelMonthlyModifiedHours.month)
            .all()
            )
    else:
        result = (
            db.session.query(
                ModelMonthlyModifiedHours.month.label('date'),
                func.sum(ModelMonthlyModifiedHours.total_hours).label('hours')
            )
            .group_by(ModelMonthlyModifiedHours.month)
            .order_by(ModelMonthlyModifiedHours.month)
            .all()
            )
    
   

       

    # Convert the result into a list of dictionaries
    dates = [row.date for row in result]
    hours = [row.hours for row in result]

    if cumulative:
        cumhours = []
        cumsum=0
        for hour in hours:
            cumsum += hour
            cumhours.append(cumsum)
        hours=cumhours

    if(unit=='TDE'):
        TDEs=[round(row/140,2) for row in hours]
        hours=TDEs
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")


    return {
        'unit':'h',
        'xAxis': {"data":dates},
        'series':[
            {
                "data": hours,
                "name": "Spent",
                "type": 'line'
            }
        ],
        'legend': {
            'data': ["Spent"]
        }
    }

#3.8. Business KPIs, pie
@app.get("/api/business_kpi/pie/hours_by_domain")
async def businesskpi_piedomainhours(
    unit: str, 
    month1: Optional[int] = None,
    year1: Optional[int] = None,
    month2: Optional[int] = None,
    year2: Optional[int] = None,
    user: SchemaHoursUserBase = Depends(get_user)):

    start_date = None
    end_date = None

    if month1 is not None and year1 is not None:
        start_date = datetime.date(year1, month1, 1)

    if month2 is not None and year2 is not None:
        if month2 == 12:
            end_date = datetime.date(year2 + 1, 1, 1)
        else:
            end_date = datetime.date(year2, month2 + 1, 1)

    if start_date:
        if end_date:
            result = (db.session.query(ModelMonthlyModifiedHours.domain, func.sum(ModelMonthlyModifiedHours.total_hours))
                .filter(
                ModelMonthlyModifiedHours.month >= start_date,
                ModelMonthlyModifiedHours.month < end_date)
                .group_by(ModelMonthlyModifiedHours.domain).all())
        else:
            result = (db.session.query(ModelMonthlyModifiedHours.domain, func.sum(ModelMonthlyModifiedHours.total_hours))
                .filter(ModelMonthlyModifiedHours.month >= start_date)
                .group_by(ModelMonthlyModifiedHours.domain).all())
    elif end_date:
        result = (db.session.query(ModelMonthlyModifiedHours.domain, func.sum(ModelMonthlyModifiedHours.total_hours))
                .filter(ModelMonthlyModifiedHours.month < end_date)
                .group_by(ModelMonthlyModifiedHours.domain).all())
    else:
        result = (db.session.query(ModelMonthlyModifiedHours.domain, func.sum(ModelMonthlyModifiedHours.total_hours))
                .group_by(ModelMonthlyModifiedHours.domain).all())

    

    factor = 1
    if(unit=='TDE'):
        factor=140
    else:
        if not unit=='h':
            raise HTTPException(status_code=401, detail="Invalid units")

    declared_hours_by_domain = {domain: round(declared_hours/factor,2) for domain, declared_hours in result}

    return {
        'unit':'h',
        'series':[
            {
                "data": declared_hours_by_domain,
                "name": "Spent hours by domain",
                "type": 'pie'
            }
        ]
    }

#3.9. Get users
@app.get("/api/getusers")
async def get_all_users(user: SchemaHoursUserBase = Depends(get_user)):
    ans=[]
    users =db.session.query(ModelHoursUser).all()
    if not users:
            raise HTTPException(status_code=404,detail="No users on database")
    
    for user in users:
        db_user={
            "name": user.username,
            "id": user.id,
            "email":user.email,
            "status": user.status}
        
        ans.append(db_user)
    return ans

#3.10. Export monthly hours
@app.get("/api/export_monthly")
async def export_monthly(month:int, year:int):
    dateM=datetime.date(year,month,1)

    temp_filename='./temp/exportModified.csv'

    query = db.session.query(ModelMonthlyModifiedHours.total_hours.label('buffer'),
                        ModelProject.project_code.label('project_code'),
                        ModelHoursUser.username.label('name'),
                        ModelMonthlyModifiedHours.month.label('month'),
                        ModelMonthlyModifiedHours.domain.label('domain'),
                        ModelMonthlyModifiedHours.total_hours.label('hours'))\
    .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id,isouter=True)\
    .join(ModelHoursUser, ModelHoursUser.id == ModelMonthlyModifiedHours.user_id,isouter=True)\
    .filter(ModelMonthlyModifiedHours.month == dateM)\
    .filter(ModelMonthlyModifiedHours.total_hours != 0)\
        .order_by(ModelProject.project_code,ModelHoursUser.username)

    if not query:
        raise HTTPException(status_code=401, detail="No records for inputted month")

    column_names = ["project_code",
                    "name",
                    "month",
                    "year",
                    "hours",
                    "domain"]

    csv_data = csv.DictWriter(open(temp_filename,"w"), fieldnames=column_names,delimiter=";")
    csv_data.writeheader()
    
    for row in query:
       
        week=row.month.isocalendar()[1]
        year=row.month.year
        
        row_data = {"project_code":row.project_code,
            "name":row.name,
            "month":week,
            "year":year,
            "hours":row.hours}      
        csv_data.writerow(row_data)

    async with aiofiles.open(temp_filename, mode="r") as f:
        contents = await f.read()


    return FileResponse(temp_filename, filename=temp_filename)

#3.11. Export project capitalization summary
@app.get("/api/export/monthly_project_capitalization")
async def export_monthly_project_capitalization(month:int, year:int,user: SchemaHoursUserBase = Depends(get_user)):
    
    dateM=datetime.date(year,month,1)

    wb=openpyxl.Workbook()
    sheet=wb.active
    ft1 = Font(size=16,
                bold=True)
    ft2 = Font(size=16)
    ft3 = Font(color="00FFFFFF",bold=True)


    fl1 = PatternFill(start_color="00000000",end_color="00000000",fill_type="solid")

    if month < 10: month='0'+str(month)
    sheet.title = str(month)+" - "+str(year)

    a1 = sheet.cell(row = 1, column= 1)
    a1.value = "R&D TDE/CAP"
    a1.font = ft1

    #Headers
    sheet.cell(row = 1, column= 2).value= str(month)+"/"+str(year)
    sheet.cell(row = 3, column= 1).value= "Project code"
    sheet.cell(row = 3, column= 2).value= "Project name"
    sheet.cell(row = 3, column= 3).value= "Sub-category"
    sheet.cell(row = 3, column= 4).value= "TDE in hours"
    sheet.cell(row = 3, column= 5).value= "CAP"
    sheet.cell(row = 3, column= 6).value= "Cap in hours"
    sheet.cell(row = 2, column= 9).value= "Capitalization summary"
    sheet.cell(row = 3, column= 9).value= "Capitalized hours"
    sheet.cell(row = 4, column= 9).value= "Division"
    sheet.cell(row = 4, column= 10).value= "Division name"
    sheet.cell(row = 4, column= 12).value= "Capitalized hours"
    
    sheet.cell(row = 3, column= 1).font= ft3
    sheet.cell(row = 3, column= 2).font= ft3
    sheet.cell(row = 3, column= 3).font= ft3
    sheet.cell(row = 3, column= 4).font= ft3
    sheet.cell(row = 3, column= 5).font= ft3
    sheet.cell(row = 3, column= 6).font= ft3
    sheet.cell(row = 2, column= 9).font= ft1
    sheet.cell(row = 3, column= 9).font= Font(bold=True)
    sheet.cell(row = 4, column= 9).font= ft3
    sheet.cell(row = 4, column= 10).font= ft3
    sheet.cell(row = 4, column= 12).font= ft3

    sheet.cell(row = 3, column= 1).fill= fl1
    sheet.cell(row = 3, column= 2).fill= fl1
    sheet.cell(row = 3, column= 3).fill= fl1
    sheet.cell(row = 3, column= 4).fill= fl1
    sheet.cell(row = 3, column= 5).fill= fl1
    sheet.cell(row = 3, column= 6).fill= fl1
    sheet.cell(row = 4, column= 9).fill= fl1
    sheet.cell(row = 4, column= 10).fill= fl1
    sheet.cell(row = 4, column= 11).fill= fl1
    sheet.cell(row = 4, column= 12).fill= fl1

    sheet.cell(row = 1, column= 2).alignment = Alignment(horizontal="right")
    sheet.cell(row = 1, column= 2).font = ft2

    sheet.column_dimensions['A'].width = 18.64
    sheet.column_dimensions['B'].width = 43
    sheet.column_dimensions['C'].width = 18.36
    sheet.column_dimensions['D'].width = 12.91
    sheet.column_dimensions['E'].width = 8.27
    sheet.column_dimensions['F'].width = 13.18
    sheet.column_dimensions['G'].width = 8.27
    sheet.column_dimensions['H'].width = 11
    sheet.column_dimensions['I'].width = 13.45
    sheet.column_dimensions['J'].width = 28.45
    sheet.column_dimensions['K'].width = 17.91
    sheet.column_dimensions['L'].width = 17.91
    
    table_by_project= db.session.query(ModelProject.id, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month == dateM,
                ModelProject.sub_category != 'ABS')\
    .group_by(ModelProject.id)\
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())


    if not table_by_project:
        raise HTTPException(status_code=401, detail="No records for inputted month")


    row_projects=4
    
    for row in table_by_project:
        line=db.session.query(ModelProject).filter(ModelProject.id == row.id).first()
        
        if(
            (line.start_cap_date is not None and line.start_cap_date <= dateM and (line.end_cap_date is None or line.end_cap_date > dateM))
        ):
            capnum=1
        else: capnum=0

        sheet.cell(row = row_projects, column= 1).value= line.project_code 
        sheet.cell(row = row_projects, column= 2).value= line.project_name 
        sheet.cell(row = row_projects, column= 3).value= line.sub_category 
        sheet.cell(row = row_projects, column= 4).value= row.hours
        sheet.cell(row = row_projects, column= 5).value= capnum 
        sheet.cell(row = row_projects, column= 6).value= float(row.hours)*float(capnum)
        row_projects+=1
    
    cap_table_by_division= db.session.query(ModelProject.division, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month == dateM,
            or_(
                and_(
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= dateM,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > dateM))))\
    .group_by(ModelProject.division) \
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())

    divisionnames={"HOME":"Home (Except headphone)",
                   "PRO":"Pro (Except headphone)",
                   "HEADPHONE":"Headphone: Home, Pro",
                   "MOTORITIES":"All motorities"}

    row_other=5
    totsum1=0

    for row in cap_table_by_division:
        sheet.cell(row = row_other, column= 9).value=row.division
        sheet.cell(row = row_other, column= 10).value=divisionnames.get(row.division,"")
        sheet.cell(row = row_other, column= 12).value=row.hours
        row_other+=1
        totsum1+=row.hours

    sheet.cell(row = row_other, column= 9).value="Subtotal"
    sheet.cell(row = row_other, column= 9).font=Font(bold=True)

    sheet.cell(row = row_other, column= 12).value=totsum1
    sheet.cell(row = row_other, column= 12).font=Font(bold=True)


    """Part 2"""
    
    FCS_hours_query = db.session.query(func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month == dateM,
                ModelProject.project_code == 'FCS')
        
    
    ETC_hours_query = db.session.query(func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month == dateM,
                ModelProject.sub_category == 'ETC',
                ModelProject.project_code != 'FCS')
        
    
    TOT_hours_query = db.session.query(func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
        .join(ModelProject, ModelProject.id == ModelMonthlyModifiedHours.project_id)\
        .filter(ModelMonthlyModifiedHours.month == dateM,
                ModelProject.sub_category != 'ABS')

    fcs=0
    etc=0
    tot=0

    for row in FCS_hours_query:fcs=row.hours
    for row in ETC_hours_query:etc=row.hours
    for row in TOT_hours_query:tot=row.hours

    if not tot: tot = 0
    if not etc: etc = 0
    if not fcs: fcs = 0

    row_other+=3

    sheet.cell(row = row_other, column= 9).value="Non-capitalized hours"
    sheet.cell(row = row_other, column= 9).font=Font(bold=True)

    row_other+=1

    sheet.cell(row = row_other, column= 9).value="Category name"
    sheet.cell(row = row_other, column= 9).font=ft3
    sheet.cell(row = row_other, column= 9).fill=fl1

    sheet.cell(row = row_other, column= 10).fill=fl1

    sheet.cell(row = row_other, column= 11).fill=fl1

    sheet.cell(row = row_other, column= 12).value="Non-cap hours"
    sheet.cell(row = row_other, column= 12).font=ft3
    sheet.cell(row = row_other, column= 12).fill=fl1

    row_other+=1
    

    sheet.cell(row = row_other, column= 9).value="Non-capitalized projects (in Concept)"
    sheet.cell(row = row_other, column= 12).value=tot-fcs-etc-totsum1
    

    row_other+=1
    
    sheet.cell(row = row_other, column= 9).value="Engineering change/Production support"
    sheet.cell(row = row_other, column= 12).value=fcs
    

    row_other+=1

    sheet.cell(row = row_other, column= 9).value="Management/Research/Regulation/ISIT/etc"
    sheet.cell(row = row_other, column= 12).value=etc

    row_other+=1

    sheet.cell(row = row_other, column= 9).value="Subtotal"
    sheet.cell(row = row_other, column= 9).font=Font(bold=True)

    sheet.cell(row = row_other, column= 12).value=tot-totsum1
    sheet.cell(row = row_other, column= 12).font=Font(bold=True)

    row_other+=2

    sheet.cell(row = row_other, column= 9).value="Total"
    sheet.cell(row = row_other, column= 9).font=Font(bold=True)

    sheet.cell(row = row_other, column= 12).value=tot
    sheet.cell(row = row_other, column= 12).font=Font(bold=True)

    row_other+=3

    cap_table_by_subcategory= db.session.query(ModelProject.sub_category, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month == dateM,
            or_(
                and_(
                    ModelProject.start_cap_date.isnot(None),
                    ModelProject.start_cap_date <= dateM,
                    or_(ModelProject.end_cap_date.is_(None), ModelProject.end_cap_date > dateM))))\
    .group_by(ModelProject.sub_category) \
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())

    tot_table_by_subcategory= db.session.query(ModelProject.sub_category, func.sum(ModelMonthlyModifiedHours.total_hours).label('hours'))\
    .join(ModelMonthlyModifiedHours, ModelProject.id == ModelMonthlyModifiedHours.project_id, isouter=True)\
    .filter(ModelMonthlyModifiedHours.month == dateM,
            ModelProject.sub_category != 'ABS') \
    .group_by(ModelProject.sub_category) \
    .having(func.sum(ModelMonthlyModifiedHours.total_hours) != 0) \
    .order_by(func.sum(ModelMonthlyModifiedHours.total_hours).desc())



    sheet.cell(row = row_other, column= 9).font=ft1
    sheet.cell(row = row_other, column= 9).value="Dev effort / Cap hours by sub categories"
    row_other+=1
    sheet.cell(row = row_other, column= 9).value="Sub-category"
    sheet.cell(row = row_other, column= 9).font=ft3
    sheet.cell(row = row_other, column= 9).fill=fl1

    sheet.cell(row = row_other, column= 10).value="Category name"
    sheet.cell(row = row_other, column= 10).font=ft3
    sheet.cell(row = row_other, column= 10).fill=fl1

    sheet.cell(row = row_other, column= 11).value="Dev effort"
    sheet.cell(row = row_other, column= 11).font=ft3
    sheet.cell(row = row_other, column= 11).fill=fl1

    sheet.cell(row = row_other, column= 12).value="Capitalization"
    sheet.cell(row = row_other, column= 12).font=ft3
    sheet.cell(row = row_other, column= 12).fill=fl1

    row_other+=1
    dict_caps={row.sub_category:row.hours for row in cap_table_by_subcategory}
    
    subcatnames={"H_CI":"Home CI",
                   "H_LSP":"Home Loudspeaker",
                   "H_HP":"Home Headphone",
                   "H_AMP":"Home Amplifier",
                   "M_CAR":"Motorities Car",
                   "M_MAR":"Motorities Marine",
                   "M_OEM":"Motorities OEM",
                   "P_LSP":"Pro Loudspeaker",
                   "P_HP":"Pro Headphone",
                   "ETC":"Others"}
    


    for row in tot_table_by_subcategory:
        sheet.cell(row = row_other, column= 9).value=row.sub_category
        sheet.cell(row = row_other, column= 10).value=subcatnames.get(row.sub_category,"")
        sheet.cell(row = row_other, column= 11).value=row.hours
        sheet.cell(row = row_other, column= 12).value=dict_caps.get(row.sub_category,0)
        
        row_other+=1

    row_other+=1

    sheet.cell(row = row_other, column= 9).value="Total"
    sheet.cell(row = row_other, column= 9).font=Font(bold=True)

    sheet.cell(row = row_other, column= 11).value=tot
    sheet.cell(row = row_other, column= 11).font=Font(bold=True)

    sheet.cell(row = row_other, column= 12).value=totsum1
    sheet.cell(row = row_other, column= 12).font=Font(bold=True)

    temp_filename = "./temp/exctest.xlsx"
    wb.save(filename=temp_filename)
    return FileResponse(temp_filename, filename=temp_filename)

#___________________________________________________________________________________________

#Auxiliary functions
def project_exists(project_id: Union[int,str]) -> bool:
    project = db.session.query(ModelProject).get(project_id)
    return project is not None

def validateManagers(BM: bool, user: SchemaHoursUserBase= Depends(get_user)):
    usertype = getattr(db.session.query(ModelHoursUser).filter(
        ModelHoursUser.username == user.username,
        ModelHoursUser.email == user.email).first(),"role")

    res=False

    if not BM:
        if (usertype=="Project Manager" or usertype=="Business Manager"):
            res=True
    else:
        if (usertype=="Business Manager"):
            res=True
    return res

def find_capitalization(project_phase: int):
    res = None
    
    if not (project_phase >=0 and project_phase <=7):
        raise HTTPException(status_code=401, detail="Invalid project phase")
    
    if project_phase <= 2:
        res = False
    else:
        res = True

    return res

def week_in_month(month, record_date):
    # Get the Monday and Wednesday of the week corresponding to the record_date
    monday = record_date - timedelta(days=record_date.weekday())
    wednesday = monday + timedelta(days=2)

    # Check if both Monday and Wednesday belong to the same month as the given month
    cond = (
        (monday.year, monday.month) == (month.year, month.month) and
        (wednesday.year, wednesday.month) == (month.year, month.month)
    )

    return cond

def update_project_phase_hours(project_id):
    # Create a session to interact with the database
    try:
        # Query ProjectPhases for the specific project_id
        phases = db.session.query(ModelProjectPhase).filter(ModelProjectPhase.project_id == project_id).all()

        # Loop through each phase and calculate the summed hours from RecordProjects
        for phase in phases:
            start_date = phase.start_date
            end_date = phase.end_date

            # Query RecordProjects and sum the declared_hours within the date interval
            total_hours = db.session.query(func.sum(ModelRecordProjects.declared_hours)).filter(
                ModelRecordProjects.project_id == project_id,
                ModelRecordProjects.date_rec >= start_date,
                ModelRecordProjects.date_rec < end_date
            ).scalar()

            # Update the hours column of the ProjectPhases with the calculated total_hours
            phase.hours = total_hours

        # Commit the changes to the database
        db.session.commit()

    except Exception as e:
        raise HTTPException(status_code=404, detail="Unable to update hours")

def concatanateListswoduplicates(list1, list2):
    set1=set(list1)
    set2=set(list2)
    result_set =  set1.union(set2)
    return list(result_set)

def convert_date_to_month(input_date):
    
        output_date = input_date.strftime('%Y-%m')
        return output_date

def convert_to_monthstr(input_date: str):    
    parts = input_date.split('-')        
    return f"{parts[0]}-{parts[1]}"

def convert_to_seconddaystr(input_date):
    
    # Split the input date into parts using '-' as the separator
    parts = input_date.split('-')
    
    # Reformat the date to 'YYYY-MM' format
    output_date = f"{parts[0]}-{parts[1]}-02"
    return output_date

#add: change monthly hours from csv. Made for one-time convenience; no need for implementation.
@app.put("/api/import-csv-monthly")
async def import_csv_monthly(file: UploadFile = File(...)):
    
    contents = await file.read()
    csv_data = csv.DictReader(contents.decode("utf-8").splitlines(),delimiter=";")
    
   
    for row in csv_data:
        month = int(row['month'])
        project_code = str(row['project'])
        hours = float(row['hours'])
        user_id = int(row['user_id'])

        dateM=datetime.date(2023,month,1)
        projectid = getattr(db.session.query(ModelProject).filter(ModelProject.project_code == project_code).first(),"id")

        searched_record=db.session.query(ModelMonthlyModifiedHours).filter(
                ModelMonthlyModifiedHours.project_id==projectid,
                ModelMonthlyModifiedHours.month==dateM,
                ModelMonthlyModifiedHours.user_id==user_id).first()
            
        if not searched_record:
            db_monthlyhour = ModelMonthlyModifiedHours(
                user_id = user_id,
                project_id = projectid,
                month = dateM,
                total_hours = hours
                )
        
            db.session.add(db_monthlyhour)
        else:
            searched_record.total_hours = hours


    db.session.commit()
    return {"message": "Import successful."}



# Get monthly_report /*/
@app.get("/api/monthly_report")
async def get_monthly_report(month: int, year: int, user = Depends(get_user)):
    
    model_monthly_report =db.session.query(ModelMonthlyReport).filter(
        ModelMonthlyReport.month == f'{year}-{month}-01'
    ).first()

    if not model_monthly_report:
            raise HTTPException(status_code=404,detail="Monthly report does not exist")
    ans = SchemaMonthlyReport(
        month=model_monthly_report.month,
        closed=model_monthly_report.closed,
        sync_date=model_monthly_report.sync_date,
    )

    return ans

# Post monthly_report /*/
@app.post("/api/monthly_report")
async def get_monthly_report(monthlyReport: SchemaMonthlyReport, user = Depends(get_user)):
    
    model_monthly_report =db.session.query(ModelMonthlyReport).filter(
        ModelMonthlyReport.month == monthlyReport.month
    ).first()

    if model_monthly_report:
            raise HTTPException(status_code=400,detail="Monthly report already exists")

    model_monthly_report = ModelMonthlyReport(
        month=monthlyReport.month,
        closed=monthlyReport.closed,
        sync_date=monthlyReport.sync_date,
    )

    db.session.add(model_monthly_report)
    db.session.commit()
    return {"message": "Monthly report created successfully."}

#  Change state monthly_report
@app.put("/api/monthly_report")
async def get_monthly_report(month: int, year: int, close: bool, user = Depends(get_user)):
    

    model_monthly_report =db.session.query(ModelMonthlyReport).filter(
        ModelMonthlyReport.month == f'{year}-{month}-01'
    ).first()

    if not model_monthly_report:
            raise HTTPException(status_code=404,detail="Monthly report does not exist")

    model_monthly_report.closed = close
    db.session.commit()
    return {"message": "Monthly report state successfully changed."}

# Get monthly_report /*/
@app.get("/api/projects/monthly-info")
async def get_projects_monthly_infos(month: int, year: int, user = Depends(get_user)):
    
    model_monthly_info =db.session.query(ModelProjectMonthlyInformation).filter(
        ModelProjectMonthlyInformation.month == f'{year}-{month}-01'
    )

    if not model_monthly_info:
            raise HTTPException(status_code=404,detail="Projets monthly informations unreachable")
    ans = []
    for mmi in model_monthly_info:
        ans.append(SchemaProjectMonthlyInformation(
                project_id = mmi.project_id,
                month = mmi.month.month,
                year = mmi.month.year,
                forecast_hours = mmi.forecast_hours,
                capitalizable = mmi.capitalizable,
        ))

    return ans
